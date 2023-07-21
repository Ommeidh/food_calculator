from openpyxl import Workbook
from openpyxl import load_workbook
import os
import datetime

def add_food_data(filename, food_data):
    wb = load_workbook(filename)
    ws = wb.active

    # Write the data to the Excel sheet
    for data_row in food_data:
        # Convert Decimal values to float
        for i in range(2, len(data_row)):
            data_row[i] = float(data_row[i])

        ws.append(data_row)

    wb.save(filename)

def create_workbook(filename):
    wb = Workbook()
    ws = wb.active
    ws.append(['Date', 'Food Name', 'Calories', 'Total Fat', 'Carbohydrates', 'Protein', 'Dietary Fiber'])
    wb.save(filename)

def load_existing_workbook(filename):
    if isinstance(filename, str) and os.path.exists(filename):
        return load_workbook(filename)
    else:
        return None

def load_closest_date_workbook():
    excel_files = [file for file in os.listdir('.') if file.endswith('.xlsx')]
    if not excel_files:
        return None

    current_date = datetime.datetime.now()
    closest_date_diff = None
    closest_file = None

    for file in excel_files:
        try:
            file_date = datetime.datetime.strptime(file.replace('nutritional_log_', '').replace('.xlsx', ''), '%Y_%m_%d_%H_%M_%S')
            date_diff = abs((current_date - file_date).total_seconds())
            if closest_date_diff is None or date_diff < closest_date_diff:
                closest_date_diff = date_diff
                closest_file = file
        except ValueError:
            pass  # Skip files with invalid date formats

    if closest_file:
        return load_workbook(closest_file)
    else:
        return None
